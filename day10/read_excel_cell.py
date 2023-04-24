# from openpyxl import load_workbook
#
# wb = load_workbook('p1.xlsx')
# # 获取地0个sheet
# sheet = wb.worksheets[0]
# # get (n,n) cell value  the local is begin with 1
# cell = sheet.cell(1, 1)
# print(cell.value)
# # 获取样式
# print(cell.style)
# # 获取字符样式
# print(cell.font)
# # 排列情况
# print(cell.alignment)
#
# # 获取某个单元格,通过sheet索引
# c1 = sheet['A2']
# print(c1.value)
#
# # 获取第n行单元格
# print(sheet[1])
# for cell in sheet[1]:
#     print(cell.value)
#
#
# # 获取所有行的第一列数据
# for row in sheet.rows:
#     print(row[0].value,row[1].value)
#
#
# # 获取每一列数据
# for col in sheet.columns:
#     print(col[1].value)


from openpyxl import load_workbook
wb = load_workbook('p1.xlsx')
sheet = wb.worksheets[2]

# 获取n行n列的单元格
c1 = sheet.cell(1, 1)
print(c1)
c2 = sheet.cell(1, 2)
print(c2)

wb2=load_workbook('p1.xlsx')
sheet=wb.worksheets[2]
for row in sheet.rows:
    print(row)