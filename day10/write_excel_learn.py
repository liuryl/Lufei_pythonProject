# from openpyxl import load_workbook
#
# wb = load_workbook('p1.xlsx')
# sheet_1 = wb.worksheets[0]
#
# # find the cell and change the value
# cell_1 = sheet_1.cell(1, 1)
# cell_1.value = 'new start'
# #
# # save to p2
# wb.save('p1.xlsx')
#
# # build a new excel to write
# from openpyxl import load_workbook
# from openpyxl import workbook
#
# # buile a new excel and build a sheet
# wb = workbook.Workbook()
# sheet_1 = wb.worksheets[0]
# # find the cell and change
# cell_1 = sheet_1.cell(1, 1)
# cell_1.value = 'new_start'
#
# # save the sheet
# wb.save('/Users/liuriyilang/PycharmProjects/Lufei_pythonProject/day10/p2.xlsx')

#
# from openpyxl import workbook
#
# wb = workbook.Workbook()
#
# # 修改sheet名字
# sheet_1 = wb.worksheets[0]
# sheet_1.title = 'data'
# # wb.save('/Users/liuriyilang/PycharmProjects/Lufei_pythonProject/day10/p2.xlsx')
#
# # 创建sheet并设置颜色
# sheet_2 = wb.create_sheet('work project', 1)
# sheet_2.sheet_properties.tabColor = '1072BA'
# wb.save('/Users/liuriyilang/PycharmProjects/Lufei_pythonProject/day10/p2.xlsx')
#
# # 打开excel时默认打开的sheet
# wb.active=1
# wb.save('p2.xlsx')
#
#
# # 拷贝sheet
# sheet_3=wb.create_sheet('work_work_project')
# sheet_3.sheet_properties.tabColor='1072BA'
# # sheet_3=wb['work_work_project']
#
# new_sheet=wb.copy_worksheet(wb['sheet_3'])
# new_sheet.title='new_project'
# wb.save('p2.xlsx')


from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill, GradientFill

wb = load_workbook('p1.xlsx')
sheet_1 = wb.worksheets[1]

cell_1 = sheet_1.cell(9, 2)
cell_1.border = Border(
    top=Side(style='thin', color='FFB6C1'),
    bottom=Side(style='dashed', color='FFB6C1'),
    left=Side(style='dashed', color='FFB6C1'),
    right=Side(style='dashed', color='FFB6C1'),
    diagonal=Side(style='thin', color='483D8B'),
    diagonalUp=True,
    diagonalDown=True
)

wb.save('p22.xlsx')
