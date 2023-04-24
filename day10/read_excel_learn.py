from openpyxl import load_workbook

xls_workbook = load_workbook('p1.xlsx')

# 获取sheet
print(xls_workbook.sheetnames)

# choose the sheet
sheet = xls_workbook['数据导出']
# 输出单元格内容
cell = sheet.cell(1, 2)
print(cell.value)

# 基于索引位置，选择sheet
sheet = xls_workbook.worksheets[0]
cell = sheet.cell(1, 2)
print(cell.value)

# 循环所有的sheet
for name in xls_workbook.sheetnames:
    sheet = xls_workbook[name]
    cell = sheet.cell(1, 1)
    print(cell.value)

for sheet in xls_workbook.worksheets:
    cell=sheet.cell(1,1)
    print(cell.value)

for sheet in xls_workbook:
    cell=sheet.cell(1,1)
    print(cell.value)